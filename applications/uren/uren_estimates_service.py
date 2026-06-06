"""Excel I/O for Ureninschattingen (project status + planned hours)."""

from __future__ import annotations

import os
from collections import defaultdict
from copy import copy
from datetime import datetime

from openpyxl.utils.cell import coordinate_from_string
from tkinter import messagebox

from uren_excel_service import (
    EXCEL_PATH,
    _excel_write_error,
    _find_open_com_workbook,
    _float_or_none,
    _load_workbook,
    _parse_excel_datum,
    _save_workbook,
    sync_excel_workbook_if_open,
)

SHEET_ESTIMATES = "Ureninschattingen"
TABLE_ESTIMATES = "Tabel132"
ESTIMATE_START_ROW = 6

COL_EST_DATUM = 1
COL_EST_WEEK = 2
COL_EST_OG = 3
COL_EST_PROJECT = 4
COL_EST_PLANNED = 5
COL_EST_ACTUAL = 6
COL_EST_UURSTATUS = 7
COL_EST_EINDSTATUS = 8
COL_EST_STATUS_MENU = 9
COL_EST_STATUS = 10
COL_EST_OPMERKING = 11

NUM_EST_COLS = COL_EST_OPMERKING

PROJECT_STATUSES = [
    "Wachten op akkoord",
    "In opdracht",
    "Afgerond",
    "On hold",
    "Geannuleerd",
]

DEFAULT_STATUS = "Wachten op akkoord"

ACTIVE_STATUSES = {"In opdracht", "On hold", "Wachten op akkoord"}


def last_hour_dates_by_project(hour_rows: list[dict]) -> dict[str, datetime]:
    """Meest recente urendatum per projectstring (exacte match op kolom Project)."""
    last: dict[str, datetime] = {}
    for r in hour_rows or []:
        proj = str(r.get("project") or "").strip()
        if not proj:
            continue
        d = r.get("datum")
        if not isinstance(d, datetime):
            continue
        if proj not in last or d > last[proj]:
            last[proj] = d
    return last


def sort_estimates_by_recent_hours(estimates: list[dict], hour_rows: list[dict]) -> list[dict]:
    """Nieuwste urenregel per project bovenaan; projecten zonder uren onderaan."""
    last = last_hour_dates_by_project(hour_rows)

    def sort_key(row: dict) -> tuple:
        proj = str(row.get("project") or "").strip()
        if proj in last:
            return (0, -last[proj].timestamp())
        est_dt = row.get("datum")
        if isinstance(est_dt, datetime):
            return (1, -est_dt.timestamp())
        return (2, 0.0)

    return sorted(estimates, key=sort_key)


def display_delta(row: dict) -> float | None:
    status = row.get("status") or DEFAULT_STATUS
    if status == "In opdracht" and row.get("uurstatus") is not None:
        return float(row["uurstatus"])
    if status == "Afgerond" and row.get("uur_eindstatus") is not None:
        return float(row["uur_eindstatus"])
    if status == "In opdracht":
        return float(row.get("ureninschatting") or 0) - float(row.get("gemaakte_uren") or 0)
    return None


def _is_estimate_row(values) -> bool:
    if not values or len(values) < COL_EST_PROJECT:
        return False
    proj = values[COL_EST_PROJECT - 1]
    return bool(str(proj or "").strip())


def _row_is_empty_estimate_slot(values) -> bool:
    if not values:
        return True
    return not str(values[COL_EST_PROJECT - 1] or "").strip()


def _estimate_table_bounds(ws):
    if TABLE_ESTIMATES not in ws.tables:
        return None
    tbl = ws.tables[TABLE_ESTIMATES]
    start_cell, end_cell = tbl.ref.split(":")
    start_col_letter, start_row = coordinate_from_string(start_cell)
    end_col_letter, end_row = coordinate_from_string(end_cell)
    return tbl, start_col_letter, start_row, end_col_letter, end_row


def _com_estimates_worksheet(wb_com):
    return wb_com.Worksheets(SHEET_ESTIMATES)


def _com_estimates_list_object(ws):
    return ws.ListObjects(TABLE_ESTIMATES)


def _com_estimates_data_bounds(ws):
    lo = _com_estimates_list_object(ws)
    body = lo.DataBodyRange
    if body is None:
        return None, None
    start = int(body.Row)
    end = int(body.Row + body.Rows.Count - 1)
    return start, end


def _com_find_add_estimate_row(ws):
    lo = _com_estimates_list_object(ws)
    table_first = int(lo.HeaderRowRange.Row) + 1
    table_last = int(lo.Range.Row + lo.Range.Rows.Count - 1)
    for row_index in range(table_first, table_last + 1):
        proj = ws.Cells(row_index, COL_EST_PROJECT).Value
        if _row_is_empty_estimate_slot((None, None, None, proj)):
            return row_index, False
    return table_last + 1, True


def _com_patch_estimate_cells(ws, row_num, fields):
    if fields.get("datum") is not None:
        ws.Cells(row_num, COL_EST_DATUM).Value = fields["datum"]
    if "opdrachtgever" in fields:
        ws.Cells(row_num, COL_EST_OG).Value = fields["opdrachtgever"]
    if "project" in fields:
        ws.Cells(row_num, COL_EST_PROJECT).Value = fields["project"]
    if "ureninschatting" in fields:
        ws.Cells(row_num, COL_EST_PLANNED).Value = fields["ureninschatting"]
    if "status" in fields:
        ws.Cells(row_num, COL_EST_STATUS).Value = fields["status"]
    if "opmerking" in fields:
        ws.Cells(row_num, COL_EST_OPMERKING).Value = fields["opmerking"]


def _find_add_row_openpyxl_estimates(ws):
    bounds = _estimate_table_bounds(ws)
    if bounds is None:
        return None
    tbl, start_col_letter, start_row, end_col_letter, end_row = bounds
    for row_index in range(start_row, end_row + 1):
        if _row_is_empty_estimate_slot(
            tuple(ws.cell(row_index, c).value for c in range(1, NUM_EST_COLS + 1))
        ):
            return row_index, False, tbl, start_col_letter, start_row, end_col_letter, end_row
    insert_row = end_row + 1
    return insert_row, True, tbl, start_col_letter, start_row, end_col_letter, end_row


def _excel_row_to_estimate(row, row_index=None):
    proj = str(row[COL_EST_PROJECT - 1] or "").strip()
    if not proj:
        return None
    dt = _parse_excel_datum(row[COL_EST_DATUM - 1])
    planned = _float_or_none(row[COL_EST_PLANNED - 1]) or 0.0
    actual = _float_or_none(row[COL_EST_ACTUAL - 1]) or 0.0
    uurstatus = _float_or_none(row[COL_EST_UURSTATUS - 1])
    eindstatus = _float_or_none(row[COL_EST_EINDSTATUS - 1])
    status = str(row[COL_EST_STATUS - 1] or "").strip() or DEFAULT_STATUS
    entry = {
        "datum": dt,
        "opdrachtgever": str(row[COL_EST_OG - 1] or "").strip(),
        "project": proj,
        "ureninschatting": planned,
        "gemaakte_uren": actual,
        "uurstatus": uurstatus,
        "uur_eindstatus": eindstatus,
        "status": status,
        "opmerking": str(row[COL_EST_OPMERKING - 1] or "").strip(),
    }
    if row_index is not None:
        entry["row_index"] = row_index
    return entry


def _iter_estimate_rows_from_com(wb_com):
    ws = _com_estimates_worksheet(wb_com)
    lo = _com_estimates_list_object(ws)
    body = lo.DataBodyRange
    if body is None:
        return
    for r in range(1, int(body.Rows.Count) + 1):
        row_num = int(body.Rows(r).Row)
        values = tuple(body.Cells(r, c).Value for c in range(1, NUM_EST_COLS + 1))
        if _is_estimate_row(values):
            yield row_num, values


def _iter_estimate_sheet_rows():
    wb_com = _find_open_com_workbook(EXCEL_PATH)
    if wb_com is not None:
        try:
            return list(_iter_estimate_rows_from_com(wb_com))
        except Exception:
            pass
    wb = _load_workbook(EXCEL_PATH, data_only=True)
    if SHEET_ESTIMATES not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[SHEET_ESTIMATES]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=ESTIMATE_START_ROW, values_only=True), start=ESTIMATE_START_ROW):
        if _is_estimate_row(row):
            rows.append((row_idx, row))
    wb.close()
    return rows


def load_estimate_rows():
    if not os.path.exists(EXCEL_PATH):
        return []
    sync_excel_workbook_if_open(EXCEL_PATH)
    out = []
    try:
        for row_idx, row in _iter_estimate_sheet_rows():
            entry = _excel_row_to_estimate(row, row_idx)
            if entry:
                out.append(entry)
    except Exception:
        return []
    return out


def summarize_by_status(rows):
    counts = defaultdict(int)
    planned_active = 0.0
    actual_active = 0.0
    remaining_active = 0.0
    over_budget = []
    for r in rows:
        st = r.get("status") or DEFAULT_STATUS
        counts[st] += 1
        if st not in ACTIVE_STATUSES:
            continue
        planned = float(r.get("ureninschatting") or 0)
        actual = float(r.get("gemaakte_uren") or 0)
        planned_active += planned
        actual_active += actual
        delta = r.get("uurstatus")
        if delta is None and st == "In opdracht":
            delta = planned - actual
        if delta is not None:
            remaining_active += float(delta)
            if float(delta) < 0:
                over_budget.append(r)
    return {
        "counts": dict(counts),
        "active_planned": round(planned_active, 2),
        "active_actual": round(actual_active, 2),
        "active_remaining": round(remaining_active, 2),
        "over_budget": over_budget,
    }


def _parse_fields(datum_str, opdrachtgever, project, ureninschatting, status, opmerking):
    datum_dt = datetime.strptime(datum_str, "%Y-%m-%d")
    planned = float(ureninschatting or 0)
    st = (status or DEFAULT_STATUS).strip()
    if st not in PROJECT_STATUSES:
        st = DEFAULT_STATUS
    return {
        "datum": datum_dt,
        "opdrachtgever": (opdrachtgever or "").strip(),
        "project": (project or "").strip(),
        "ureninschatting": planned,
        "status": st,
        "opmerking": (opmerking or "").strip(),
    }


def add_estimate_row(datum_str, opdrachtgever, project, ureninschatting, status, opmerking=""):
    if not (project or "").strip():
        messagebox.showerror("Fout", "Project is verplicht.")
        return False
    fields = _parse_fields(datum_str, opdrachtgever, project, ureninschatting, status, opmerking)

    wb_com = _find_open_com_workbook(EXCEL_PATH)
    if wb_com is not None:
        try:
            ws = _com_estimates_worksheet(wb_com)
            row_num, need_insert = _com_find_add_estimate_row(ws)
            if need_insert:
                ws.Rows(row_num).Insert()
            _com_patch_estimate_cells(ws, row_num, fields)
            wb_com.Save()
            return True
        except Exception as exc:
            _excel_write_error("Fout bij project toevoegen", exc)
            return False

    wb = _load_workbook(EXCEL_PATH)
    if SHEET_ESTIMATES not in wb.sheetnames:
        wb.close()
        messagebox.showerror("Fout", f"Tabblad {SHEET_ESTIMATES} niet gevonden.")
        return False
    ws = wb[SHEET_ESTIMATES]
    add_info = _find_add_row_openpyxl_estimates(ws)
    if add_info is None:
        wb.close()
        messagebox.showerror("Fout", f"Tabel {TABLE_ESTIMATES} niet gevonden.")
        return False
    insert_row, need_insert, tbl, start_col_letter, start_row, end_col_letter, end_row = add_info
    if need_insert:
        ws.insert_rows(insert_row)
        last_row = insert_row - 1
        if last_row >= start_row:
            for col in range(1, NUM_EST_COLS + 1):
                source_cell = ws.cell(row=last_row, column=col)
                target_cell = ws.cell(row=insert_row, column=col)
                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)
        tbl.ref = f"{start_col_letter}{start_row}:{end_col_letter}{end_row + 1}"

    ws.cell(row=insert_row, column=COL_EST_DATUM, value=fields["datum"])
    ws.cell(row=insert_row, column=COL_EST_OG, value=fields["opdrachtgever"])
    ws.cell(row=insert_row, column=COL_EST_PROJECT, value=fields["project"])
    ws.cell(row=insert_row, column=COL_EST_PLANNED, value=fields["ureninschatting"])
    ws.cell(row=insert_row, column=COL_EST_STATUS, value=fields["status"])
    ws.cell(row=insert_row, column=COL_EST_OPMERKING, value=fields["opmerking"])
    _save_workbook(wb, EXCEL_PATH)
    wb.close()
    return True


def update_estimate_row(row_index, datum_str, opdrachtgever, project, ureninschatting, status, opmerking=""):
    if not (project or "").strip():
        messagebox.showerror("Fout", "Project is verplicht.")
        return False
    fields = _parse_fields(datum_str, opdrachtgever, project, ureninschatting, status, opmerking)

    wb_com = _find_open_com_workbook(EXCEL_PATH)
    if wb_com is not None:
        try:
            ws = _com_estimates_worksheet(wb_com)
            start_row, end_row = _com_estimates_data_bounds(ws)
            if start_row is None or row_index < start_row or row_index > end_row:
                messagebox.showerror("Fout", "Projectrij niet meer gevonden (ververs lijst).")
                return False
            _com_patch_estimate_cells(ws, row_index, fields)
            wb_com.Save()
            return True
        except Exception as exc:
            _excel_write_error("Fout bij project bewerken", exc)
            return False

    wb = _load_workbook(EXCEL_PATH)
    ws = wb[SHEET_ESTIMATES]
    bounds = _estimate_table_bounds(ws)
    if bounds is None:
        wb.close()
        messagebox.showerror("Fout", f"Tabel {TABLE_ESTIMATES} niet gevonden.")
        return False
    _, _, start_row, _, end_row = bounds
    if row_index < start_row or row_index > end_row:
        wb.close()
        messagebox.showerror("Fout", "Projectrij niet meer gevonden (ververs lijst).")
        return False
    ws.cell(row=row_index, column=COL_EST_DATUM, value=fields["datum"])
    ws.cell(row=row_index, column=COL_EST_OG, value=fields["opdrachtgever"])
    ws.cell(row=row_index, column=COL_EST_PROJECT, value=fields["project"])
    ws.cell(row=row_index, column=COL_EST_PLANNED, value=fields["ureninschatting"])
    ws.cell(row=row_index, column=COL_EST_STATUS, value=fields["status"])
    ws.cell(row=row_index, column=COL_EST_OPMERKING, value=fields["opmerking"])
    _save_workbook(wb, EXCEL_PATH)
    wb.close()
    return True


def delete_estimate_row(row_index):
    wb_com = _find_open_com_workbook(EXCEL_PATH)
    if wb_com is not None:
        try:
            ws = _com_estimates_worksheet(wb_com)
            start_row, end_row = _com_estimates_data_bounds(ws)
            if start_row is None or row_index < start_row or row_index > end_row:
                messagebox.showerror("Fout", "Projectrij niet meer gevonden (ververs lijst).")
                return False
            ws.Rows(row_index).Delete()
            wb_com.Save()
            return True
        except Exception as exc:
            _excel_write_error("Fout bij project verwijderen", exc)
            return False

    wb = _load_workbook(EXCEL_PATH)
    ws = wb[SHEET_ESTIMATES]
    bounds = _estimate_table_bounds(ws)
    if bounds is None:
        wb.close()
        messagebox.showerror("Fout", f"Tabel {TABLE_ESTIMATES} niet gevonden.")
        return False
    tbl, start_col_letter, start_row, end_col_letter, end_row = bounds
    if row_index < start_row or row_index > end_row:
        wb.close()
        messagebox.showerror("Fout", "Projectrij niet meer gevonden (ververs lijst).")
        return False
    ws.delete_rows(row_index, 1)
    tbl.ref = f"{start_col_letter}{start_row}:{end_col_letter}{end_row - 1}"
    _save_workbook(wb, EXCEL_PATH)
    wb.close()
    return True
